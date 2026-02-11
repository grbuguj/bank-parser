import io
from typing import List
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from models.transaction import Transaction


def create_excel(transactions: List[Transaction], bank_name: str) -> bytes:
    """Transaction 리스트를 엑셀 파일로 변환 후 바이트 반환"""
    wb = Workbook()
    ws = wb.active
    ws.title = bank_name[:31]

    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F4F8F", end_color="2F4F8F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_alignment_center = Alignment(horizontal="center", vertical="center")
    data_alignment_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 헤더
    headers = ["거래은행", "입금일", "출금일", "금액", "거래사유"]
    col_widths = [15, 18, 18, 18, 45]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # 데이터 행 (거래은행 컬럼은 빈칸, 아래서 병합 처리)
    for row_idx, t in enumerate(transactions, start=2):
        row_data = [
            "",
            t.deposit_date,
            t.withdraw_date,
            t.amount,
            t.reason,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            if col_idx == 4:  # 금액
                cell.number_format = "#,##0"
                cell.alignment = data_alignment_center
            elif col_idx == 5:  # 거래사유
                cell.alignment = data_alignment_left
            else:
                cell.alignment = data_alignment_center

        ws.row_dimensions[row_idx].height = 20

    # 거래은행 전체 병합
    total_rows = len(transactions)
    if total_rows == 1:
        ws.cell(row=2, column=1, value=bank_name)
    elif total_rows > 1:
        ws.merge_cells(start_row=2, start_column=1, end_row=total_rows + 1, end_column=1)
        ws.cell(row=2, column=1, value=bank_name)

    if total_rows >= 1:
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # 바이트로 저장
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
